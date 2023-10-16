import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
from handle_code import location_handle,dowload_handle,trends_handle,post_handle
import openpyxl
from tkinter import filedialog
import os
import concurrent.futures
file_Path = ""
import threading
def create_home_frame(container):
    save_media_var = tk.IntVar()
    home_frame = tk.Frame(container)

    # Tạo LabelFrame với văn bản "Lấy các hashtag"
    hashtag_frame = ttk.LabelFrame(home_frame, text="Lấy các hashtag")
    hashtag_frame.place(relx=0, rely=0, relwidth=0.2, relheight=1)

    # Tạo nút "Lấy các hashtag"
    def change_loc():
        if location_handle.change_location():
            messagebox.showinfo("Thông báo", "Đã đổi vị trí")
        else:
            messagebox.showinfo("Thông báo", "Đổi vị trí thất bại")
        
        
    Location_button = tk.Button(hashtag_frame, text="Đổi location", command=change_loc)
    Location_button.pack(fill="x")
    get_hashtag_button = tk.Button(hashtag_frame, text="Lấy các Hashtag Trending", command=lambda: add_to_listbox())
    get_hashtag_button.pack(fill="x")

    # Tạo Listbox
    listbox = tk.Listbox(hashtag_frame, font=("Arial", 10))
    listbox.pack(fill="both", expand=True)
    def delete_item_listBox(event):
        selected_item = listbox.get(listbox.nearest(event.y))  # Lấy item được chọn khi chuột phải
        confirmed = messagebox.askokcancel("Xác nhận", f"Xóa '{selected_item}'?")
        if confirmed:
            listbox.delete(listbox.nearest(event.y))  # Xóa item được chọn
    
    listbox.bind("<Delete>", delete_item_listBox)
    def add_to_listbox():
        List = trends_handle.get_Trends()
        for tag in List:
                if tag[0]:
                    if tag[0] not in listbox.get(0, "end"):
                        listbox.insert("end", tag[0])

    def delete_all_data():
        result = messagebox.askquestion("Xác nhận", "Bạn có chắc chắn muốn xóa tất cả dữ liệu của danh sách?")
        if result == "yes":
            listbox.delete(0, "end")
    
    delete_button = tk.Button(hashtag_frame, text="Xóa danh sách Hashtag", command=delete_all_data)
    delete_button.pack(fill="x")

    # Tạo LabelFrame mới bên phải hashtag_frame
    posts_frame = ttk.LabelFrame(home_frame, text="Lấy các bài viết")
    posts_frame.place(relx=0.2, rely=0, relwidth=1, relheight=0.2)

    # Thêm mô tả và TextBox nhập số từ 1 đến 100
    label_description = ttk.Label(posts_frame, text="Số lượng bài viết lấy(n): ")
    label_description.grid(row=0, column=0, padx=10, pady=5, sticky="w")

    entry_number = ttk.Entry(posts_frame, font=("Arial", 10), width=40)
    entry_number.grid(row=0, column=1, padx=10, pady=5, sticky="w")

    # Thêm TextBox và label mô tả
    label_hashtag = ttk.Label(posts_frame, text="Hashtag: ")
    label_hashtag.grid(row=1, column=0, padx=10, pady=5, sticky="w")

    entry_hashtag = ttk.Entry(posts_frame, font=("Arial", 10), width=40)
    entry_hashtag.grid(row=1, column=1, padx=10, pady=5, sticky="w")
    
    # Gán giá trị mặc định và đặt focus vào entry_number
    entry_number.insert(0, "10")
    entry_number.focus_set()
    # Định nghĩa font với kích thước 14
    custom_font = ("Arial", 12)
    def on_mouse_scroll(event):
        delta = -event.delta  # Lấy giá trị scroll delta và đảo ngược nó để đảm bảo hướng cuộn đúng
        table.yview_scroll(delta, "units")
    # Thiết lập font cho mục trong listbox
    listbox.configure(font=custom_font)
    def on_listbox_select(event):
        selected_item = listbox.get(listbox.curselection())  # Lấy item được chọn trong Listbox
        entry_hashtag.delete(0, "end")  # Xóa giá trị hiện tại của entry_hashtag
        entry_hashtag.insert(0, selected_item)  # Gán giá trị của item vào entry_hashtag

    listbox.bind("<<ListboxSelect>>", on_listbox_select)
    def start_button_cmd():
        hashtag_text = entry_hashtag.get()
        number_text = entry_number.get()
        if not hashtag_text or not number_text:
            messagebox.showinfo("Thông báo", f"Không được bỏ trống các trường")
        else:
            List = post_handle.Get_n_Post_By_Input_Trend(hashtag_text,int(number_text))
            add_data_to_table(List)
    start_button = tk.Button(posts_frame, text="Lấy n bài viết của Hashtag đã chọn", command=start_button_cmd)
    start_button.grid(row=2, column=0, padx=1, pady=1)
    # Button to "Lấy n bài viết"
    def process_trend(tag, number_text):
        global List_post
        List_post.extend(post_handle.Get_n_Post_By_Input_Trend(tag, number_text))
    def get_n_posts_button_cmd():
        global List_post
        List_post = []
        number_text = entry_number.get()
        if not number_text:
            messagebox.showinfo("Thông báo", f"Nhập số lượng bài viết muốn lấy")
        else:   
            max_threads = 5
            with concurrent.futures.ThreadPoolExecutor(max_threads) as executor:
                for i in range(listbox.size()):
                    j = i
                    tag = listbox.get(j)
                    executor.submit(process_trend, tag, int(number_text))
            add_data_to_table(List_post)
    get_n_posts_button = tk.Button(posts_frame, text="Lấy n bài viết của mỗi Hashtag", command=get_n_posts_button_cmd)
    get_n_posts_button.grid(row=2, column=1, padx=1, pady=1)

    def add_data_to_listbox():
        hashtag_text = entry_hashtag.get()
        if hashtag_text not in listbox.get(0, "end") and hashtag_text !="":
            listbox.insert("end", hashtag_text)
            entry_hashtag.delete(0, "end")
    add_hashtag_button = tk.Button(posts_frame, text="Thêm hashtag vào danh sách", command=add_data_to_listbox)
    add_hashtag_button.grid(row=2, column=2, padx=1, pady=1)
    # Hàm kiểm tra xem giá trị nhập vào có hợp lệ hay không
    def is_valid_number(value):
        try:
            num = int(value)
            return 1 <= num <= 1000
        except ValueError:
            return False

    # Hàm để kiểm tra giá trị khi TextBox mất focus
    def validate_entry(event):
        value = entry_number.get()
        if not is_valid_number(value):
            entry_number.delete(0, "end")
            entry_number.insert(0, "10")
            error_label.config(text="Chỉ có thể lấy từ 1 đến 1000 bài viết!", foreground="red")
        else:
            error_label.config(text="", foreground="black")

    entry_number.bind("<FocusOut>", validate_entry)
    entry_hashtag.bind("<Return>", lambda event=None: add_data_to_listbox())
    # Label để hiển thị thông báo lỗi
    error_label = ttk.Label(posts_frame, text="", font=("Arial", 11))
    error_label.grid(row=3, column=0, columnspan=2, padx=1, pady=1, sticky="w")

    # Create a LabelFrame for the table below posts_frame
    table_frame = ttk.LabelFrame(home_frame, text="Table")
    table_frame.place(relx=0.2, rely=0.2, relwidth=0.8, relheight=0.8)

    # Create a Treeview widget for the table
    columns = ("Name","User", "Status", "Tags", "Interactions", "Link","Tag")
    table = ttk.Treeview(table_frame, columns=columns, show="headings")

    table.pack(fill="both", expand=True)

    # Add column headings
    for col in columns:
        table.heading(col, text=col)
        table.column(col, width=100)  # Set the column width for other columns
        # Function to add data to the table
    def add_data_to_table(List):
        for item in List:
            name = item[1]
            user = item[2]
            status = item[3]
            tags = ', '.join(item[4])
            link = item[0]
            interactions = "RL: "+item[5]+" RL: "+item[6]," LK: "+item[7]+" W: "+item[8]
            table.insert("", "end", values=(name,user, status, tags,interactions, link,item[9]))
    def add_tags_from_selected_item():
        selected_item = table.selection()[0]
        if selected_item:
            tags = table.item(selected_item, "values")[3]
            if tags == "None": return
            tags = tags.split(", ")  # Tách các thẻ từ chuỗi
            for tag in tags:
                if tag:
                    if tag not in listbox.get(0, "end"):
                        listbox.insert("end", tag)
    # Function to clear data from the table
    def clear_table_data():
        result = messagebox.askyesno("Thông báo", "Bạn muốn xóa tất cả data trên table?")
        if result:
            for item in table.get_children():
                table.delete(item)
    def delete_selected_data():
        selected_items = table.selection()
        if not selected_items:
            messagebox.showinfo("Thông báo", "Vui lòng chọn hàng để xóa.")
        else:
            confirm = messagebox.askokcancel("Xác nhận", "Bạn có muốn xóa hàng đã chọn?")
            if confirm:
                for item in selected_items:
                    table.delete(item)
    table.bind('<Delete>', lambda event=None: delete_selected_data())
    def export_to_excel_with_dialog():
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")])

        if file_path:
            workbook = openpyxl.Workbook()
            sheet = workbook.active

            # Write the column headings to the Excel sheet
            for i, col in enumerate(columns):
                sheet.cell(row=1, column=i+1, value=col)

            # Write the data from the table to the Excel sheet
            for i, row in enumerate(table.get_children()):
                for j, col in enumerate(columns):
                    value = table.item(row, "values")[j]
                    sheet.cell(row=i+2, column=j+1, value=value)

            # Save the Excel file to the selected location
            workbook.save(file_path)
            
            # Extract the directory path
            folder_path = "/".join(file_path.split("/")[:-1])
            

            # Create "Photo" and "Video" subfolders
            photo_folder = os.path.join(folder_path, "Photo")
            video_folder = os.path.join(folder_path, "Video")

            os.makedirs(photo_folder, exist_ok=True)
            os.makedirs(video_folder, exist_ok=True)
            
            # Open the folder in File Explorer (works on Windows)
            link_value =[]
            name_value = []
            if save_media_var.get() == 1:
                for item in table.get_children():
                    link_value.append( table.item(item, "values")[5] )
                    name_value.append( table.item(item, "values")[0])
                dowload_handle.dowload_Videos_Or_Photos(link_value,name_value,folder_path)
            messagebox.showinfo("Thông báo", f"Dữ liệu đã được xuất ra tệp Excel.\nThư mục lưu: {folder_path}")
            os.startfile(folder_path)
    save_media_var = tk.IntVar()
    def import_from_excel():
        file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])  # Open a file dialog to choose an Excel file

        if file_path:
            try:
                workbook = openpyxl.load_workbook(file_path)
                sheet = workbook.active

                # Assuming that the Excel columns match the table's columns, you can directly read the data
                data = []
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    data.append(row)

                for item in data:
                    table.insert("", "end", values=item)
                messagebox.showinfo("Thông báo", f"Dữ liệu đã được nhập từ tệp Excel:\n{file_path}")

            except Exception as e:
                messagebox.showerror("Lỗi", f"Có lỗi xảy ra khi nhập dữ liệu từ tệp Excel: {str(e)}")
    
    # Button to clear data from the table
    add_tag_button = tk.Button(table_frame, text="Thêm hashtag vào danh sách", command=add_tags_from_selected_item)
    add_tag_button.pack(side="left", padx=10, pady=5)
    clear_data_button = tk.Button(table_frame, text="Xóa bảng", command=clear_table_data)
    clear_data_button.pack(side="left", padx=10, pady=5)
    delete_data_button = tk.Button(table_frame, text="Xóa hàng", command=delete_selected_data)
    delete_data_button.pack(side="left", padx=10, pady=5)
    export_excel_button = tk.Button(table_frame, text="Lưu File Excel", command=export_to_excel_with_dialog)
    export_excel_button.pack(side="left", padx=10, pady=5)
    save_media_var.set(1)
    save_media_checkbox = ttk.Checkbutton(table_frame, text="Lưu kèm hình ảnh và video", variable=save_media_var, onvalue=1, offvalue=0)
    save_media_checkbox.pack(side="left", padx=10, pady=5)
    import_excel_button = tk.Button(table_frame, text="Nhập File Excel", command=import_from_excel)
    import_excel_button.pack(side="left", padx=10, pady=5)
    table.bind("<MouseWheel>", on_mouse_scroll)
    return home_frame
